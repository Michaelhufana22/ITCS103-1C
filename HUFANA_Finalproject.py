import tkinter as tk
from tkinter import ttk, messagebox
import openpyxl as op


FILE = ("HUFANA_Database.xlsx")


def display():
    workbook = op.load_workbook(FILE)
    sheet = workbook.active

    # Clear table
    for row in table.get_children():
        table.delete(row)

    # Display records
    for row in sheet.iter_rows(min_row=2, values_only=True):
        table.insert("", tk.END, values=row)

def validate():
    if not name_entry.get():
        messagebox.showerror("Error", "Name is required!")
        return False

    if not room_entry.get():
        messagebox.showerror("Error", "Room is required!")
        return False

    if not contact_entry.get():
        messagebox.showerror("Error", "Contact is required!")
        return False

    if not contact_entry.get().isdigit():
        messagebox.showerror("Error", "Contact must be numbers")
        return False

    if not movein_entry.get():
        messagebox.showerror("Error", "Move-In Date is required!")
        return False

    if not rent_entry.get():
        messagebox.showerror("Error", "Rent is required!")
        return False

    if not rent_entry.get().isdigit():
        messagebox.showerror("Error", "Rent must be numeric!")
        return False
    
    rent_value = int(rent_entry.get())
    if rent_value > 5000 :
        messagebox.showerror("Error","Maximum rent is 5000 only")

    elif rent_value < 3000 :
        messagebox.showerror("Youre eligible for bedspace only")
    if not status_combo.get():
        messagebox.showerror("Error", "Select payment status!")
        return False

    if rent_value == 5000 :
        status_combo.set("Fully Paid",)
        status_combo.config(state="readonly")
    return True


def add():
    if not validate():
        return

    workbook = op.load_workbook(FILE)
    sheet = workbook.active

    new_id = sheet.max_row

    sheet.append([new_id,name_entry.get(),room_entry.get(),contact_entry.get(),movein_entry.get(),rent_entry.get(),
        status_combo.get()])

    workbook.save(FILE)

    messagebox.showinfo("Success", "Tenant added successfully!")

    clear()
    display()

def auto_populate(event):
    selected = table.focus()

    if not selected:
        return

    values = table.item(selected, "values")

    clear()

    name_entry.insert(0, values[1])
    room_entry.insert(0, values[2])
    contact_entry.insert(0, values[3])
    movein_entry.insert(0, values[4])
    rent_entry.insert(0, values[5])
    status_combo.set(values[6])

def update():
    selected = table.focus()

    if not selected:
        messagebox.showerror("Error", "Select a record first!")
        return

    if not validate():
        return

    values = table.item(selected, "values")
    record_id = values[0]

    workbook = op.load_workbook(FILE)
    sheet = workbook.active

    for row in sheet.iter_rows(min_row=2):
        if str(row[0].value) == str(record_id):
            row[1].value = name_entry.get()
            row[2].value = room_entry.get()
            row[3].value = contact_entry.get()
            row[4].value = movein_entry.get()
            row[5].value = rent_entry.get()
            row[6].value = status_combo.get()
            break

    workbook.save(FILE)

    messagebox.showinfo("Success", "Record updated successfully!")
    
    clear()


def delete():
    selected = table.focus()

    if not selected:
        messagebox.showerror("Error", "Select a record first!")
        return

    confirm = messagebox.askyesno("Confirm Delete","Are you sure you want to delete this record?")

    if not confirm:
        return

    values = table.item(selected, "values")
    record_id = values[0]

    workbook = op.load_workbook(FILE)
    sheet = workbook.active

    for i, row in enumerate(sheet.iter_rows(min_row=2), start=2):
        if str(row[0].value) == str(record_id):
            sheet.delete_rows(i)
            break

    workbook.save(FILE)

    messagebox.showinfo("Success", "Record deleted successfully!")
    display()
    clear()


def clear():
    name_entry.delete(0, tk.END)
    room_entry.delete(0, tk.END)
    contact_entry.delete(0, tk.END)
    movein_entry.delete(0, tk.END)
    rent_entry.delete(0, tk.END)
    status_combo.set("")


window = tk.Tk()
window.title("Boarding House Management System")
window.geometry("1000x500")
window.configure(bg="lightblue")

title = tk.Label(window,text="Boarding House Management System",font=("Arial", 14, "bold"),bg="white"
)
title.grid(row=0, column=0, columnspan=4, pady=10)

frame = tk.Frame(window, bg="lightblue", bd=5, relief="ridge")
frame.grid(row=1, column=0, columnspan=4, padx=10, pady=10)

tk.Label(frame, text="Name", bg="lightblue").grid(row=0, column=0, padx=5, pady=5)
name_entry = tk.Entry(frame, width=30)
name_entry.grid(row=0, column=1)

tk.Label(frame, text="Room", bg="lightblue").grid(row=1, column=0, padx=5, pady=5)
room_entry = tk.Entry(frame, width=30)
room_entry.grid(row=1, column=1)

tk.Label(frame, text="Contact", bg="lightblue").grid(row=2, column=0, padx=5, pady=5)
contact_entry = tk.Entry(frame, width=30)
contact_entry.grid(row=2, column=1)

tk.Label(frame, text="Move-In Date", bg="lightblue").grid(row=3, column=0, padx=5, pady=5)
movein_entry = tk.Entry(frame, width=30)
movein_entry.grid(row=3, column=1)

tk.Label(frame, text="Rent", bg="lightblue").grid(row=4, column=0, padx=5, pady=5)
rent_entry = tk.Entry(frame, width=30)
rent_entry.grid(row=4, column=1)

tk.Label(frame, text="Status", bg="lightblue").grid(row=5, column=0, padx=5, pady=5)

status_combo = ttk.Combobox(
    frame,values=["Half paid", "Fully paid"],state="readonly",width=27)
status_combo.grid(row=5, column=1)

tk.Button(window,text="Add",bg="green",fg="white",width=12,command=add
).grid(row=2, column=0, pady=10)

tk.Button(window,text="Update",bg="orange",fg="white",width=12,command=update
).grid(row=2, column=1)

tk.Button(window,text="Delete",bg="red",fg="white",width=12,command=delete).grid(row=2, column=2)

tk.Button(window,text="Clear",width=12,command=clear).grid(row=2, column=3)


columns = ("ID","Name","Room","Contact","Move-In","Rent","Status")

table = ttk.Treeview(window,columns=columns,show="headings",height=12)

for col in columns:
    table.heading(col, text=col)
    table.column(col, width=130)

table.grid(row=3,column=0,columnspan=4,padx=10,pady=10)

table.bind("<<TreeviewSelect>>", auto_populate)

display()

window.mainloop()